from django.shortcuts import render,redirect
from django.views import View
from django.contrib.auth import logout
import pytz
from django.utils import timezone
from django.contrib import messages
from adminside.repositories.all_repository import DocumentRepository, SchoolYearRepository, SectionRepository, StudentListHistoryRepository
from adminside.services.all_service import SectionService, StudentListHistoryService
from landingpage.models import StudentInformation, Section, EnrollmentManagement
from django.contrib.auth.decorators import user_passes_test, login_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse
import json
from django.http import HttpResponse
from openpyxl import load_workbook
import os
from django.conf import settings
from django.http import HttpResponse
import logging
from io import BytesIO


logger = logging.getLogger(__name__)

def is_teacher(user):
    return user.is_authenticated and user.user_role in ["Teacher"]



# STUDENT LIST VIEW FOR TEACHER
@method_decorator(
    [ login_required(login_url="/authentication/sign-in/"), user_passes_test(is_teacher) ],
    name="dispatch",
)
class StudentListView(View):
    def get(self, request):
        # get the teacher logged in
        teacher_info = getattr(request.user, "teacherinformation", None)
        if not teacher_info:
            return redirect("signin")
        documents = DocumentRepository.get_all()
        school_year = SchoolYearRepository.get_all().order_by("updated_at").last()
        section = SectionRepository.filter(teacher=teacher_info, academic_year=school_year).first()
        return render(request, "teacher/index.html", { "section": section, "school_year": school_year, "teacher_info": teacher_info, "documents": documents})

# STUDENT LIST API FOR TEACHER
@method_decorator(
    [ login_required(login_url="/authentication/sign-in/"), user_passes_test(is_teacher) ],
    name="dispatch",
)
class StudentListAPI(View):
    def get(self, request):
        teacher_info = getattr(request.user, "teacherinformation", None)
        if not teacher_info:
            return redirect("signin")
        school_year = SchoolYearRepository.get_all().order_by("updated_at").last()
        section = SectionRepository.filter(teacher=teacher_info, academic_year=school_year).first()
        response_data = StudentListHistoryService.get_student_list_history_for_datatables(request, teacher_info.grade_level, school_year, section.section_name)
        return JsonResponse(response_data, safe=False)
    
# INPUT FINAL AVERAGE
@method_decorator(
    [ login_required(login_url="/authentication/sign-in/"), user_passes_test(is_teacher) ],
    name="dispatch",
)
class InputFinalAverageView(View):
    def post(self, request):
        response = StudentListHistoryService.input_final_average(request)
        return JsonResponse(response, safe=False)

# MARK SECTION AS COMPLETED
@method_decorator(
    [ login_required(login_url="/authentication/sign-in/"), user_passes_test(is_teacher) ],
    name="dispatch",
)
class MarkAsCompletedSectionView(View):
    def post(self, request):
        school_year = SchoolYearRepository.get_all().order_by("updated_at").last()
        response = SectionService.mark_as_completed_section(request, school_year)
        return JsonResponse(response, safe=False)

@method_decorator(
    [ login_required(login_url="/authentication/sign-in/"), user_passes_test(is_teacher) ],
    name="dispatch",
)
class ExportExcelFile(View):
    def get(self, request):
        teacher_info = getattr(request.user, "teacherinformation", None)
        if not teacher_info:
            return redirect("signin")

        # === Get latest school year ===
        school_year = SchoolYearRepository.get_all().order_by("updated_at").last()
        if not school_year:
            return HttpResponse("No school year found.", status=404)

        # === Get the section by teacher ===
        section = SectionRepository.filter(
            teacher=teacher_info, academic_year=school_year
        ).first()
        if not section:
            return HttpResponse("No section found for this teacher.", status=404)

        # === Get student list ordered by last name
        student_list_history = StudentListHistoryRepository.filter(
            grade_level=teacher_info.grade_level,
            section=section,
            school_year=school_year,
            student_information__student_status="Enrolled",
            student_information__user__deactivated=False,
        ).select_related("student_information", "section", "teacher_information").order_by("student_information__last_name")

        # === Load Excel template ===
        template_path = os.path.join(
            settings.BASE_DIR, "enrollmentwebsite/static/excel/Masterlist.xlsx"
        )
        wb = load_workbook(template_path)
        ws = wb["INPUT DATA"]

        # === Fill Header Information ===
        ws.cell(row=5, column=7, value="PASO DE BLAS NATIONAL HIGH SCHOOL")  # G5:R5
        ws.cell(row=5, column=33, value=school_year.name)                    # AG5:AJ5
        ws.cell(row=7, column=11, value=f"{teacher_info.grade_level} - {section.section_name}")  # K7:P7
        ws.cell(row=7, column=19, value=f"{teacher_info.first_name} {teacher_info.last_name}")   # S7:AB7

        # === Fill Learner Data ===
        start_row = 9
        name_col = 2   # B9:E9
        gender_col = 6  # F9:P9
        lrn_col = 17    # Q9:AB9
        birthdate_col = 29  # AC9:AJ9

        for i, student in enumerate(student_list_history, start=1):
            row = start_row + (i - 1)

            s_info = student.student_information
            student_name = f"{s_info.last_name}, {s_info.first_name} {s_info.middle_name or ''}".strip()
            gender = s_info.gender or ""
            lrn = s_info.lrn or ""
            birthdate = s_info.birth_date.strftime("%B %d, %Y") if s_info.birth_date else ""

            # Write only in the top-left cell of merged ranges
            ws.cell(row=row, column=name_col, value=student_name)
            ws.cell(row=row, column=gender_col, value=gender)
            ws.cell(row=row, column=lrn_col, value=lrn)
            ws.cell(row=row, column=birthdate_col, value=birthdate)

        # === Export ===
        filename = f"Masterlist_{section.section_name}_{school_year.name}.xlsx"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
        

@method_decorator(
    [
        login_required(login_url="/authentication/sign-in/"),
        user_passes_test(is_teacher),
    ],
    name="dispatch",
)
class ExportStudentsExcelView(View):
    def get(self, request):
        try:
            # Get teacher's section
            teacher_info = getattr(request.user, "teacherinformation", None)
            section = Section.objects.filter(teacher=teacher_info).first()

            if not section:
                return HttpResponse("Section not found.", status=404)

            students = StudentInformation.objects.filter(
                section=section.section_name,
                student_status="Enrolled",
                user__deactivated=False,
                school_year=section.academic_year,
            ).order_by("last_name")

            # Load Excel template
            template_path = os.path.join(
                settings.BASE_DIR, "enrollmentwebsite/static/excel/Masterlist.xlsx"
            )
            if not os.path.exists(template_path):
                return HttpResponse("Template file not found.", status=404)

            wb = load_workbook(template_path)
            ws = wb.active

            # Start rows
            male_row = 12
            female_row = 63

            for student in students:
                full_name = f"{student.last_name}, {student.first_name} {student.middle_name or ''}".strip()
                if student.gender.lower() == "male":
                    ws[f"B{male_row}"] = full_name
                    male_row += 1
                elif student.gender.lower() == "female":
                    ws[f"B{female_row}"] = full_name
                    female_row += 1

            # Prepare response
            filename = f"Masterlist_{section.section_name}.xlsx"
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response

        except Exception as e:
            return HttpResponse(f"Error: {str(e)}", status=500)


class TeacherLogoutView(View):
    def get(self, request):
        if request.user.is_authenticated:
            user = request.user
            philippine_time = timezone.now().astimezone(pytz.timezone("Asia/Manila"))
            user.is_active = False
            user.updated_at = philippine_time
            user.save(update_fields=["is_active", "updated_at"])
            request.session.flush()

        logout(request)
        messages.success(request, "Logged out successfully!")

        return redirect("signin")


# Create your views here.
